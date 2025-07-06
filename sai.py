import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinterdnd2 import TkinterDnD, DND_FILES
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import logging

def setup_logging(output_dir):
    """
    Set up logging to a file in the output directory.
    """
    log_file = os.path.join(output_dir, 'debug.log')
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

def is_invalid_description(description):
    """
    Determine if a description is invalid based on criteria:
    - Contains only shopping platforms (Amazon.com, Amazon.ca, Walmart.com, eBay, etc.)
    - Contains primarily URLs (including short links like a.co) with minimal content
    - Contains error messages (Error:, Failed, etc.)
    - Lacks meaningful product information
    """
    if pd.isna(description) or description.strip() == "":
        logging.debug(f"Invalid: Empty or NaN description")
        return True

    # Normalize description for checking
    desc = description.strip()
    desc_lower = desc.lower()

    # List of shopping platforms
    shopping_platforms = [
        'amazon.com', 'amazon.ca', 'walmart.com', 'ebay', 'google image result',
        'verify your identity'
    ]

    # Extract status and clean description
    status_pattern = r'^(new open box|used like new|used|sealed|opened)\s*[-–—]?\s*'
    clean_desc = re.sub(status_pattern, '', desc_lower, flags=re.IGNORECASE).strip()

    # Check if description is just a platform name
    if clean_desc in shopping_platforms:
        logging.debug(f"Invalid: Description is only a platform name - {desc}")
        return True

    # Enhanced URL detection
    url_pattern = r'(https?://\S+|www\.\S+|a\.co/\S+|bit\.ly/\S+|tinyurl\.com/\S+)'
    urls = re.findall(url_pattern, clean_desc)
    
    if urls:
        # Remove URLs and check remaining content
        non_url_content = re.sub(url_pattern, '', clean_desc).strip()
        non_url_words = non_url_content.split()
        
        # If description is just a URL or has minimal additional content
        if not non_url_content or len(non_url_words) <= 3:
            logging.debug(f"Invalid: Primarily URL with minimal content - {desc}")
            return True
        
        # Check if remaining content is just a platform name or trivial
        if any(platform in non_url_content for platform in shopping_platforms) and len(non_url_words) <= 5:
            logging.debug(f"Invalid: URL with platform name and minimal content - {desc}")
            return True

    # Check for error messages
    error_keywords = ['error:', 'failed', 'product title not found']
    if any(keyword in clean_desc for keyword in error_keywords):
        logging.debug(f"Invalid: Contains error message - {desc}")
        return True

    # Check if description lacks meaningful content
    if len(clean_desc.split()) <= 3 and any(platform in clean_desc for platform in shopping_platforms):
        logging.debug(f"Invalid: Minimal content with platform - {desc}")
        return True

    logging.debug(f"Valid: {desc}")
    return False

def format_status(description):
    """
    Format the description to have the status in uppercase, followed by ' - ', then the rest.
    """
    if pd.isna(description) or description.strip() == "":
        return description

    # Extract status using regex
    status_pattern = r'^(new open box|used like new|used|sealed|opened)\s*[-–—]?\s*(.*)'
    match = re.match(status_pattern, description, re.IGNORECASE)
    if match:
        status, rest = match.groups()
        return f"{status.upper()} - {rest.strip()}"
    else:
        # If no status found, check if description starts with a status
        for status in ['new open box', 'used like new', 'used', 'sealed', 'opened']:
            if description.lower().startswith(status):
                return f"{status.upper()} - {description[len(status):].strip()}"
        return description  # Return unchanged if no status identified

def format_excel_file(df, filename):
    """
    Save DataFrame to Excel with proper formatting:
    - Create a table
    - Autofit column widths
    - Apply text wrapping
    - Use professional styling
    """
    # Save DataFrame to Excel
    df.to_excel(filename, index=False, engine='openpyxl')

    # Load workbook for formatting
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Define table range
    max_row = ws.max_row
    max_col = ws.max_column
    table_range = f"A1:{chr(64 + max_col)}{max_row}"

    # Create table
    tab = Table(displayName="DataTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    cell_font = Font(size=11)
    alignment = Alignment(wrap_text=True, vertical='top')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply styles and autofit columns
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = header_font if cell.row == 1 else cell_font
            cell.alignment = alignment
            cell.border = thin_border
            # Estimate column width based on content
            if cell.value:
                max_length = max(
                    len(str(line)) for line in str(cell.value).split('\n')
                )
                adjusted_width = min(max_length + 2, 100)  # Cap width at 100
                col_letter = cell.column_letter
                if ws.column_dimensions[col_letter].width is None or ws.column_dimensions[col_letter].width < adjusted_width:
                    ws.column_dimensions[col_letter].width = adjusted_width

    # Set row height for better readability
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 15 if row == 1 else 30

    wb.save(filename)

def process_excel_file(filepath):
    """
    Process the Excel file:
    - Update third column format
    - Separate valid and invalid rows
    - Save to formatted Excel files
    """
    try:
        # Set up logging
        output_dir = os.path.dirname(filepath)
        setup_logging(output_dir)

        # Read Excel file
        df = pd.read_excel(filepath)

        # Verify columns
        if len(df.columns) < 3:
            messagebox.showerror("Error", "Excel file must have at least 3 columns.")
            return

        # Update third column format
        df.iloc[:, 1] = df.iloc[:, 1].apply(format_status)
        df.iloc[:, 2] = df.iloc[:, 2].apply(format_status)

        # Identify invalid rows
        df['is_invalid'] = df.iloc[:, 2].apply(is_invalid_description)

        # Separate valid and invalid DataFrames
        valid_df = df[~df['is_invalid']].drop(columns=['is_invalid'])
        invalid_df = df[df['is_invalid']].drop(columns=['is_invalid'])

        # Define output paths
        valid_path = os.path.join(output_dir, 'valid.xlsx')
        invalid_path = os.path.join(output_dir, 'invalid.xlsx')

        # Save and format Excel files
        if not valid_df.empty:
            format_excel_file(valid_df, valid_path)
        else:
            logging.warning("No valid rows found.")
        if not invalid_df.empty:
            format_excel_file(invalid_df, invalid_path)
        else:
            logging.warning("No invalid rows found.")

        messagebox.showinfo(
            "Success",
            f"Processing complete!\nValid rows saved to: {valid_path}\nInvalid rows saved to: {invalid_path}\nCheck debug.log for details."
        )

    except Exception as e:
        logging.error(f"Error processing file: {str(e)}")
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def on_drop(event):
    """
    Handle drag-and-drop event
    """
    filepath = event.data.strip('{}')  # Remove curly braces from dropped file path
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        label.config(text=f"Processing: {os.path.basename(filepath)}")
        process_excel_file(filepath)
        label.config(text="Drag and drop an Excel file here")
    else:
        messagebox.showerror("Error", "Please drop a valid Excel file (.xlsx or .xls)")

def browse_file():
    """
    Open file dialog to select Excel file
    """
    filepath = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if filepath:
        label.config(text=f"Processing: {os.path.basename(filepath)}")
        process_excel_file(filepath)
        label.config(text="Drag and drop an Excel file here")

# Set up GUI
root = TkinterDnD.Tk()
root.title("Excel File Processor")
root.geometry("600x400")

# Create and configure drag-and-drop area
frame = tk.Frame(root, bg="#f0f0f0", relief="sunken", bd=2)
frame.pack(expand=True, fill="both", padx=20, pady=20)

label = tk.Label(
    frame,
    text="Drag and drop an Excel file here",
    font=("Arial", 14),
    bg="#f0f0f0",
    wraplength=500
)
label.pack(expand=True)

# Bind drag-and-drop event
frame.drop_target_register(DND_FILES)
frame.dnd_bind('<<Drop>>', on_drop)

# Add browse button with white background, black text, and a border for visibility
browse_button = tk.Button(
    root,
    text="Browse File",
    command=browse_file,
    font=("Arial", 12),
    bg="#FFFFFF",
    fg="#000000",
    relief="solid",
    bd=1,
    padx=10,
    pady=5
)
browse_button.pack(pady=10)

# Start the GUI
root.mainloop()