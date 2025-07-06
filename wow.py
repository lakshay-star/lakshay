import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
import requests
from bs4 import BeautifulSoup
import threading
import queue
from datetime import datetime
import logging
from fake_useragent import UserAgent
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import chardet
import time
from urllib.parse import urlparse, urlunparse
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='combined_extractor.log',
    filemode='a'
)

# Mapping of short status to full-form descriptions
STATUS_MAP = {
    "nob": "New Open Box",
    "uln": "Used Like New",
    "sealed": "Sealed",
    "opened": "Opened",
    "used": "Used"
}

class ProductExtractor:
    def __init__(self):
        self.request_timeout = 30  # Increased timeout
        self.max_retries = 3
        self.retry_delay = 2  # Seconds between retries
        self.ua = UserAgent()
        self.session = self._create_session()
        
    def _create_session(self):
        """Create a configured session with retry strategy"""
        session = requests.Session()
        
        # Configure retry strategy
        retry_strategy = Retry(
            total=self.max_retries,
            backoff_factor=1,
            status_forcelist=[408, 429, 500, 502, 503, 504]
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        
        session.headers.update({
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.5",
            "Accept-Encoding": "gzip, deflate",
        })
        
        return session

    def clean_url(self, url):
        """Clean and normalize URL"""
        try:
            parsed = urlparse(url)
            if not parsed.scheme:
                parsed = parsed._replace(scheme='https')
            if not parsed.netloc:
                return None
            # Remove fragments and query parameters that might cause issues
            cleaned = parsed._replace(fragment='', query='')
            return urlunparse(cleaned)
        except Exception:
            return None

    def clean_amazon_title(self, title):
        """Clean Amazon product title by removing unwanted prefixes and suffixes"""
        if not title:
            return ""
            
        # Remove Amazon.com prefixes and variations
        title = re.sub(r'^Amazon\.(com|\w{2,3})\s*[:|-]\s*', '', title, flags=re.IGNORECASE)
        title = re.sub(r'^Amazon\s*[:|-]\s*', '', title, flags=re.IGNORECASE)
        
        # Remove trailing Amazon references
        title = re.sub(r'\s*:\s*Amazon\.(com|\w{2,3})\s*$', '', title, flags=re.IGNORECASE)
        title = re.sub(r'\s*-\s*Amazon\.(com|\w{2,3})\s*$', '', title, flags=re.IGNORECASE)
        
        # Remove any remaining colons/dashes at start
        title = re.sub(r'^[:|-]\s*', '', title)
        
        # Clean up extra whitespace and special characters
        title = ' '.join(title.split())
        title = re.sub(r'[^\w\s-]', '', title)  # Remove special chars except spaces and hyphens
        
        return title.strip()[:500]  # Limit length

    def get_product_title(self, url):
        """Fetch and clean product title from URL"""
        cleaned_url = self.clean_url(url)
        if not cleaned_url:
            return "Invalid URL"
            
        for attempt in range(self.max_retries):
            try:
                self.session.headers.update({"User-Agent": self.ua.random})
                response = self.session.get(
                    cleaned_url, 
                    timeout=self.request_timeout,
                    allow_redirects=True
                )
                response.raise_for_status()
                
                # Detect encoding from response headers and content
                if response.encoding:
                    encoding = response.encoding
                else:
                    encoding = chardet.detect(response.content)['encoding']
                
                soup = BeautifulSoup(response.content.decode(encoding, errors='replace'), 'html.parser')
                title = None
                
                # Try multiple title extraction methods with priority order
                title_selectors = [
                    {"id": "productTitle"},
                    {"id": "title"},
                    {"class": "a-size-large product-title-word-break"},
                    {"data-feature-name": "title"},
                    {"class": "a-size-extra-large"},
                    {"class": "a-size-medium a-color-base a-text-normal"},
                    {"class": "product-title"},  # Common alternative class
                    {"itemprop": "name"},  # Schema.org itemprop
                ]
                
                for selector in title_selectors:
                    title_element = soup.find(**selector)
                    if title_element:
                        title = title_element.get_text(strip=True)
                        if title:  # Only break if we got a non-empty title
                            break
                
                if not title:
                    # Fallback to meta tags
                    meta_title = soup.find("meta", property="og:title") or soup.find("meta", attrs={"name": "title"})
                    if meta_title:
                        title = meta_title.get("content", "").strip()
                    else:
                        title_tag = soup.find("title")
                        if title_tag:
                            title = title_tag.get_text(strip=True)
                
                if title:
                    return self.clean_amazon_title(title)
                
                return "Product Title Not Found"
                
            except requests.exceptions.RequestException as e:
                logging.warning(f"Request failed (attempt {attempt + 1}): {str(e)}")
                if attempt < self.max_retries - 1:
                    time.sleep(self.retry_delay)
                else:
                    return f"Error: Failed to retrieve product title ({str(e)})"
            except Exception as e:
                logging.error(f"Error fetching title: {str(e)}")
                return f"Error: {str(e)}"
        
        return "Error: Max retries reached"

class CombinedExtractorApp:
    def __init__(self, root):
        self.root = root
        self.product_extractor = ProductExtractor()
        self.queue = queue.Queue()
        self.running = False
        self.current_file = None
        self.setup_ui()
        self.check_queue()
        self.center_window()

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')

    def setup_ui(self):
        self.root.title("WhatsApp Lot & Product Extractor")
        self.root.geometry("600x350")
        self.root.config(bg="#f5f5f5")
        self.root.resizable(False, False)
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TFrame', background="#f5f5f5")
        style.configure('TLabel', background="#f5f5f5", font=('Segoe UI', 9))
        style.configure('TButton', font=('Segoe UI', 9))
        style.configure('TProgressbar', thickness=20)
        
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(expand=True, fill=tk.BOTH)
        
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            header_frame,
            text="WhatsApp Lot & Product Extractor",
            font=('Segoe UI', 14, 'bold'),
            foreground="#333"
        ).pack(side=tk.LEFT)
        
        drop_frame = ttk.Frame(main_frame, relief=tk.SOLID, borderwidth=1)
        drop_frame.pack(expand=True, fill=tk.BOTH)
        
        self.drop_label = ttk.Label(
            drop_frame,
            text="Drag and drop your WhatsApp .txt file here\nor click to browse",
            justify=tk.CENTER,
            font=('Segoe UI', 11),
            foreground="#666",
            wraplength=400
        )
        self.drop_label.pack(expand=True, padx=20, pady=40)
        
        self.progress = ttk.Progressbar(main_frame, orient=tk.HORIZONTAL, length=400, mode='determinate')
        self.progress.pack(pady=(15, 5))
        
        self.status_label = ttk.Label(main_frame, text="Ready to process files", foreground="#555")
        self.status_label.pack()
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=(15, 0))
        
        ttk.Button(btn_frame, text="Browse File", command=self.browse_file).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Exit", command=self.root.quit).pack(side=tk.LEFT, padx=5)
        
        # Enable drag and drop functionality
        try:
            from tkinterdnd2 import DND_FILES, TkinterDnD
            drop_frame.drop_target_register(DND_FILES)
            drop_frame.dnd_bind('<<Drop>>', self.on_drop)
            self.dnd_available = True
        except ImportError:
            self.dnd_available = False
        
        drop_frame.bind("<Button-1>", self.browse_file)

    def browse_file(self, event=None):
        if self.running:
            messagebox.showwarning("Already Running", "Please wait for the current process to complete")
            return
            
        filepath = filedialog.askopenfilename(
            title="Select WhatsApp Text File",
            filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
        )
        if filepath:
            self.process_file(filepath)

    def on_drop(self, event):
        if not self.dnd_available or self.running:
            return
            
        filepath = event.data.strip('{}')
        if os.path.isfile(filepath) and filepath.lower().endswith('.txt'):
            self.process_file(filepath)
        else:
            messagebox.showerror("Invalid File", "Please drop a valid .txt file")

    def process_file(self, filepath):
        if self.running:
            messagebox.showwarning("Already Running", "Please wait for the current process to complete")
            return
            
        self.running = True
        self.current_file = filepath
        self.progress["value"] = 0
        self.status_label.config(text="Processing...", foreground="#333")
        
        threading.Thread(
            target=self._process_file_thread,
            args=(filepath,),
            daemon=True
        ).start()

    def detect_file_encoding(self, filepath):
        """Detect file encoding using chardet with larger sample size"""
        try:
            with open(filepath, 'rb') as f:
                rawdata = f.read(50000)  # Read first 50KB to detect encoding more accurately
                result = chardet.detect(rawdata)
                confidence_threshold = 0.8  # Only trust high confidence detections
                
                if result['confidence'] > confidence_threshold:
                    return result['encoding']
                else:
                    # Try common encodings if detection confidence is low
                    for encoding in ['utf-8', 'utf-16', 'iso-8859-1', 'windows-1252']:
                        try:
                            with open(filepath, 'r', encoding=encoding) as test_file:
                                test_file.read(1024)  # Test read
                                return encoding
                        except UnicodeDecodeError:
                            continue
                    
                    return 'utf-8'  # Final fallback
        except Exception as e:
            logging.warning(f"Error detecting encoding: {str(e)}")
            return 'utf-8'  # Fallback to utf-8

    def extract_data(self, text):
        """Extract product data from WhatsApp text with improved pattern matching"""
        # Normalize line endings and remove any special characters that might break parsing
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Enhanced regex pattern to handle different WhatsApp message formats
        # Handles both phone numbers and names as senders with better date/time handling
        message_pattern = r"""
            (?P<date>\d{1,2}[/-]\d{1,2}[/-]\d{2,4}),?\s*  # Date part with optional comma
            (?P<time>\d{1,2}:\d{2}\s*[ap]m)\s*-\s*  # Time part
            (?:  # Sender part (either phone number or name)
                (?:\+\d[\d\s-]{6,}\d)|  # Phone number format with international code
                (?:[^:]+)  # Name format
            ):\s*
            (?P<message>.*?)  # Message content
            (?=\n\d{1,2}[/-]\d{1,2}[/-]\d{2,4},?\s*\d{1,2}:\d{2}\s*[ap]m|$)  # Lookahead for next message or end
        """
        
        entries = []
        for match in re.finditer(message_pattern, text, re.VERBOSE | re.DOTALL | re.IGNORECASE):
            entries.append(match.group('message').strip())
        
        data = []
        link_title_cache = {}  # Cache to store URL-title mappings
        total_links = sum(1 for entry in entries if re.search(r"https?://", entry))
        processed_links = 0

        for entry in entries:
            if not entry.strip():
                continue
                
            # Find lot number (handles variations like "Lot", "lot", "LOT", "Lot No", etc.)
            lot_match = re.search(r"(?:lot|Lot|LOT)\s*(?:no|No|NO)?\s*[:#]?\s*(\d+)", entry, re.IGNORECASE)
            if not lot_match:
                continue
            lot_number = lot_match.group(1)

            # Find status keyword (case insensitive) with word boundaries
            status_match = re.search(r"\b(nob|uln|sealed|opened|used|new open box|used like new)\b", entry, re.IGNORECASE)
            if not status_match:
                continue
            status = status_match.group(1).lower()
            # Normalize status to standard form
            if status in ['new open box']:
                status = 'nob'
            elif status in ['used like new']:
                status = 'uln'
            full_form = STATUS_MAP.get(status, status.capitalize())

            # Find the first valid link (handles URLs that might be split across lines)
            link_match = re.search(r"(https?://[^\s]+)", entry)
            if not link_match:
                continue
            raw_link = link_match.group(1).strip()
            link = self.product_extractor.clean_url(raw_link)
            if not link:
                continue

            # Get product title (from cache or fetch new)
            if link in link_title_cache:
                product_title = link_title_cache[link]
            else:
                product_title = self.product_extractor.get_product_title(link)
                link_title_cache[link] = product_title
                processed_links += 1
                progress = int((processed_links / total_links) * 90) if total_links > 0 else 0
                self.queue.put(("progress", progress))

            # Compose column 2 (Status and Product Title)
            if status == "nob":
                status_title = product_title
            else:
                status_title = f"{full_form} - {product_title}"

            # Compose column 3 (Full Description)
            # Get all lines between status and URL
            lines = [line.strip() for line in entry.splitlines() if line.strip()]
            desc_lines = []
            status_found = False
            
            for line in lines:
                if re.search(r"\b" + re.escape(status) + r"\b", line, re.IGNORECASE):
                    status_found = True
                    continue
                if status_found and not line.startswith(("http://", "https://")):
                    desc_lines.append(line)

            description = f"{full_form}"
            if desc_lines:
                description += ". " + " ".join(desc_lines)
            description += f" - {product_title}"

            data.append([lot_number, status_title, description])

        return data

    def _format_excel(self, filepath, df):
        """Apply professional formatting to the Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_font = Font(size=11)
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        alignment = Alignment(wrap_text=True, vertical='top')
        
        # Write headers
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data rows
        for row_num, row_data in enumerate(df.itertuples(index=False), 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = alignment
        
        # Set column widths
        column_widths = {
            'A': 15,  # Lot Number
            'B': 40,  # Status and Product Title
            'C': 80   # Full Description
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Apply filter
        ws.auto_filter.ref = ws.dimensions
        
        # Add conditional formatting for error messages
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("Error:", "Product Title Not Found")):
                    cell.fill = red_fill
        
        # Save the workbook
        try:
            wb.save(filepath)
        except PermissionError:
            # Try alternative filename if permission denied
            alt_path = os.path.join(
                os.path.dirname(filepath),
                f"extracted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            wb.save(alt_path)
            return alt_path
        return filepath

    def _process_file_thread(self, filepath):
        try:
            # Detect file encoding first
            encoding = self.detect_file_encoding(filepath)
            
            # Read file with detected encoding and error handling
            try:
                with open(filepath, "r", encoding=encoding, errors='replace') as f:
                    text = f.read()
            except UnicodeDecodeError:
                # Fallback to utf-8 with error replacement
                with open(filepath, "r", encoding='utf-8', errors='replace') as f:
                    text = f.read()

            self.queue.put(("status", "Extracting data from messages..."))
            data = self.extract_data(text)
            if not data:
                self.queue.put(("error", "No valid entries were found in the file."))
                return

            self.queue.put(("status", "Creating data frame..."))
            df = pd.DataFrame(data, columns=["Lot Number", "Status and Product Title", "Full Description"])
            
            # Sort by Lot Number (numeric sort)
            try:
                df['Lot Number'] = pd.to_numeric(df['Lot Number'])
                df = df.sort_values('Lot Number')
                df['Lot Number'] = df['Lot Number'].astype(str)
            except ValueError:
                # If lot numbers aren't purely numeric, sort as strings
                df = df.sort_values('Lot Number')

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = os.path.splitext(os.path.basename(filepath))[0]
            output_path = os.path.join(
                os.path.dirname(filepath),
                f"{base_name}_extracted_{timestamp}.xlsx"
            )
            
            self.queue.put(("status", "Formatting Excel file..."))
            # Save with professional formatting
            saved_path = self._format_excel(output_path, df)
            self.queue.put(("complete", saved_path))
            
        except Exception as e:
            logging.error(f"Error processing file: {str(e)}", exc_info=True)
            self.queue.put(("error", f"An error occurred: {str(e)}"))
        finally:
            self.queue.put(("done", None))

    def check_queue(self):
        try:
            while True:
                msg_type, msg_data = self.queue.get_nowait()
                
                if msg_type == "progress":
                    self.progress["value"] = msg_data
                elif msg_type == "status":
                    self.status_label.config(text=msg_data)
                elif msg_type == "complete":
                    messagebox.showinfo(
                        "Complete",
                        f"✅ Data extracted successfully!\n\nSaved to:\n{msg_data}"
                    )
                    self.status_label.config(text="Completed", foreground="green")
                    self.progress["value"] = 100
                    # Open the output directory
                    try:
                        output_dir = os.path.dirname(msg_data)
                        os.startfile(output_dir)
                    except Exception:
                        pass
                elif msg_type == "error":
                    messagebox.showerror("Error", msg_data)
                    self.status_label.config(text="Error occurred", foreground="red")
                    self.progress["value"] = 0
                elif msg_type == "done":
                    self.running = False
                    self.current_file = None
                    if self.progress["value"] < 100:
                        self.progress["value"] = 0
                    
        except queue.Empty:
            pass
        
        self.root.after(100, self.check_queue)

if __name__ == "__main__":
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()  # Use TkinterDnD if available
    except ImportError:
        root = tk.Tk()
    
    # Set window icon if available
    try:
        root.iconbitmap(default='icon.ico')  # Provide your icon file
    except Exception:
        pass
    
    app = CombinedExtractorApp(root)
    root.mainloop()